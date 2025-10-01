"""
Spreadsheet Manager for Healthcare Bill Tracking

Manages Excel spreadsheets for HSA reconciliation and tax auditing.
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class SpreadsheetManager:
    """Manages healthcare bill tracking spreadsheets."""

    def __init__(self, workbook_path: str = "healthcare_bills.xlsx"):
        """
        Initialize the spreadsheet manager.

        Args:
            workbook_path: Path to the Excel workbook file
        """
        self.workbook_path = Path(workbook_path).expanduser().resolve()
        self.df = self._load_or_create_dataframe()

    def _load_or_create_dataframe(self) -> pd.DataFrame:
        """Load existing spreadsheet or create a new one."""
        if self.workbook_path.exists():
            try:
                df = pd.read_excel(self.workbook_path, sheet_name="Bills")
                # Ensure date column is datetime
                if "date" in df.columns:
                    df["date"] = pd.to_datetime(df["date"])
                return df
            except Exception as e:
                print(f"Warning: Could not load existing file: {e}")
                return self._create_empty_dataframe()
        else:
            return self._create_empty_dataframe()

    def _create_empty_dataframe(self) -> pd.DataFrame:
        """Create an empty dataframe with the required columns."""
        return pd.DataFrame(columns=[
            "date",
            "provider",
            "amount",
            "category",
            "description",
            "pdf_path",
            "year",
            "month",
            "added_on",
        ])

    def add_record(self, record: Dict[str, Any]) -> None:
        """
        Add a new healthcare bill record.

        Args:
            record: Dictionary containing bill information
                Required keys: provider, date, amount
                Optional keys: category, description, pdf_path
        """
        # Parse date
        date_str = record.get("date", "")
        try:
            date = pd.to_datetime(date_str)
        except Exception:
            date = pd.Timestamp.now()

        # Create new row
        new_row = {
            "date": date,
            "provider": record.get("provider", ""),
            "amount": float(record.get("amount", 0.0)),
            "category": record.get("category", "medical"),
            "description": record.get("description", ""),
            "pdf_path": record.get("pdf_path", ""),
            "year": date.year,
            "month": date.month,
            "added_on": pd.Timestamp.now(),
        }

        # Add to dataframe
        self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)

    def get_summary(self) -> Dict[str, Any]:
        """
        Get a summary of all tracked bills.

        Returns:
            Dictionary containing summary statistics
        """
        if self.df.empty:
            return {
                "total_records": 0,
                "total_amount": 0.0,
                "message": "No records found"
            }

        summary = {
            "total_records": len(self.df),
            "total_amount": float(self.df["amount"].sum()),
            "date_range": {
                "earliest": str(self.df["date"].min()),
                "latest": str(self.df["date"].max()),
            },
            "by_category": self.df.groupby("category")["amount"].sum().to_dict(),
            "by_year": self.df.groupby("year")["amount"].sum().to_dict(),
            "by_provider": self.df.groupby("provider")["amount"].sum().to_dict(),
        }

        return summary

    def save(self) -> None:
        """Save the dataframe to the Excel workbook with formatting."""
        if self.df.empty:
            print("Warning: No data to save")
            return

        # Sort by date
        self.df = self.df.sort_values("date", ascending=False)

        # Create Excel writer
        with pd.ExcelWriter(self.workbook_path, engine="openpyxl") as writer:
            # Write main bills sheet
            self.df.to_excel(writer, sheet_name="Bills", index=False)

            # Write summary sheets
            self._write_summary_sheets(writer)

            # Format the workbook
            workbook = writer.book
            self._format_bills_sheet(workbook["Bills"])

        print(f"Saved to {self.workbook_path}")

    def _write_summary_sheets(self, writer: pd.ExcelWriter) -> None:
        """Write summary sheets to the workbook."""
        if self.df.empty:
            return

        # By year
        year_summary = self.df.groupby("year").agg({
            "amount": "sum",
            "provider": "count"
        }).rename(columns={"provider": "num_bills"})
        year_summary.to_excel(writer, sheet_name="By Year")

        # By category
        category_summary = self.df.groupby("category").agg({
            "amount": "sum",
            "provider": "count"
        }).rename(columns={"provider": "num_bills"})
        category_summary.to_excel(writer, sheet_name="By Category")

        # By provider
        provider_summary = self.df.groupby("provider").agg({
            "amount": "sum",
            "date": "count"
        }).rename(columns={"date": "num_bills"}).sort_values("amount", ascending=False)
        provider_summary.to_excel(writer, sheet_name="By Provider")

    def _format_bills_sheet(self, worksheet) -> None:
        """Apply formatting to the Bills sheet."""
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def export_for_taxes(self, year: int, output_path: str = None) -> str:
        """
        Export bills for a specific year in a tax-friendly format.

        Args:
            year: The tax year to export
            output_path: Output file path (default: tax_export_{year}.xlsx)

        Returns:
            Path to the exported file
        """
        if output_path is None:
            output_path = f"tax_export_{year}.xlsx"

        output_path = Path(output_path).expanduser().resolve()

        # Filter by year
        year_df = self.df[self.df["year"] == year].copy()

        if year_df.empty:
            raise ValueError(f"No records found for year {year}")

        # Sort by date
        year_df = year_df.sort_values("date")

        # Calculate totals
        total_amount = year_df["amount"].sum()

        # Create tax export with summary
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Detail sheet
            year_df[["date", "provider", "category", "amount", "description"]].to_excel(
                writer, sheet_name="Details", index=False
            )

            # Summary by category
            category_summary = year_df.groupby("category")["amount"].sum().reset_index()
            category_summary.columns = ["Category", "Total Amount"]
            category_summary.to_excel(writer, sheet_name="Category Summary", index=False)

            # Monthly summary
            year_df["month_name"] = pd.to_datetime(year_df["date"]).dt.strftime("%B")
            monthly_summary = year_df.groupby(["month", "month_name"])["amount"].sum().reset_index()
            monthly_summary = monthly_summary.sort_values("month")
            monthly_summary[["month_name", "amount"]].to_excel(
                writer, sheet_name="Monthly Summary", index=False
            )

            # Overall summary
            summary_data = {
                "Metric": ["Total Amount", "Number of Bills", "Average Bill Amount", "Tax Year"],
                "Value": [
                    f"${total_amount:.2f}",
                    len(year_df),
                    f"${year_df['amount'].mean():.2f}",
                    year
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

        print(f"Tax export for {year} saved to {output_path}")
        return str(output_path)

    def export_hsa_reconciliation(self, output_path: str = None) -> str:
        """
        Export data in HSA reconciliation format.

        Args:
            output_path: Output file path (default: hsa_reconciliation.xlsx)

        Returns:
            Path to the exported file
        """
        if output_path is None:
            output_path = "hsa_reconciliation.xlsx"

        output_path = Path(output_path).expanduser().resolve()

        # Create HSA-eligible categories
        hsa_eligible_categories = ["medical", "dental", "vision", "prescription", "pharmacy"]

        # Filter for HSA-eligible expenses
        hsa_df = self.df[self.df["category"].isin(hsa_eligible_categories)].copy()

        if hsa_df.empty:
            raise ValueError("No HSA-eligible records found")

        # Sort by date
        hsa_df = hsa_df.sort_values("date")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Main HSA sheet
            hsa_df[["date", "provider", "category", "amount", "description", "pdf_path"]].to_excel(
                writer, sheet_name="HSA Expenses", index=False
            )

            # Yearly totals
            yearly = hsa_df.groupby("year")["amount"].sum().reset_index()
            yearly.columns = ["Year", "Total HSA-Eligible Amount"]
            yearly.to_excel(writer, sheet_name="Yearly Totals", index=False)

        print(f"HSA reconciliation saved to {output_path}")
        return str(output_path)
